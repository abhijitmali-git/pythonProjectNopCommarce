import openpyxl

file="TestData/TestData_login.xlsx"
def getRowcount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return (sheet.max_row)


def Columncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return (sheet.max_column)

def readDate(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    return sheet.cell(row=rownum,column=columnno).value


def writeData(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    sheet.cell(row=rownum,column=columnno).value= data
    workbook.save(file)
